from flask import (
    Flask, render_template, request,
    redirect, url_for, session, send_file
)
import json, csv, io
from functools import wraps
from openpyxl import Workbook, load_workbook
from datetime import datetime

from services.repository import StudentRepository
from services.models import Student
from services.validators import validate_student_payload

app = Flask(__name__)
app.secret_key = "secret123"
repo = StudentRepository()

# ========== AUTH ==========
def load_users():
    with open("users.json", "r") as f:
        return json.load(f)["users"]

def login_required(f):
    @wraps(f)
    def wrap(*a, **k):
        if "user" not in session:
            return redirect("/login")
        return f(*a, **k)
    return wrap

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        for u in load_users():
            if u["username"] == request.form["username"] and u["password"] == request.form["password"]:
                session["user"] = u["username"]
                return redirect("/dashboard")
    return render_template("login.html")

@app.route("/logout")
def logout():
    session.clear()
    return redirect("/login")

# ========== DASHBOARD ==========
@app.route("/")
@login_required
def home():
    return redirect("/dashboard")

@app.route("/dashboard")
@login_required
def dashboard():
    students = repo.load()
    stats = {
        "total": len(students),
        "jurusan": len(set(s.major for s in students)),
        "avg_nim": round(sum(len(s.nim) for s in students) / len(students), 2) if students else 0,
        "last_update": datetime.now().strftime("%d %b %Y %H:%M")
    }
    return render_template("dashboard.html", students=students, stats=stats)

# ========== DATA MAHASISWA ==========
@app.route("/mahasiswa")
@login_required
def mahasiswa():
    return render_template("index.html", students=repo.load())

@app.route("/add", methods=["GET", "POST"])
@login_required
def add():
    if request.method == "POST":
        data = validate_student_payload(request.form)
        student = Student(
            _nim=data["nim"],
            _name=data["name"],
            _email=data["email"],
            _major=data["major"]
        )
        repo.upsert(student)
        return redirect("/mahasiswa")
    return render_template("form.html")

@app.route("/delete/<nim>")
@login_required
def delete(nim):
    repo.delete(nim)
    return redirect("/mahasiswa")

# ========== EXPORT ==========
@app.route("/export/csv")
@login_required
def export_csv():
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["nim","name","email","major"])
    for s in repo.load():
        writer.writerow([s.nim,s.name,s.email,s.major])
    mem = io.BytesIO(output.getvalue().encode())
    return send_file(mem, as_attachment=True, download_name="mahasiswa.csv")

@app.route("/export/excel")
@login_required
def export_excel():
    wb = Workbook()
    ws = wb.active
    ws.append(["nim","name","email","major"])
    for s in repo.load():
        ws.append([s.nim,s.name,s.email,s.major])
    mem = io.BytesIO()
    wb.save(mem)
    mem.seek(0)
    return send_file(mem, as_attachment=True, download_name="mahasiswa.xlsx")

# ========== IMPORT ==========
@app.route("/import/csv", methods=["POST"])
@login_required
def import_csv():
    reader = csv.DictReader(io.StringIO(request.files["file"].read().decode()))
    for r in reader:
        repo.upsert(Student(r["nim"], r["name"], r["email"], r["major"]))
    return redirect("/dashboard")

@app.route("/import/excel", methods=["POST"])
@login_required
def import_excel():
    wb = load_workbook(request.files["file"])
    for i, row in enumerate(wb.active.iter_rows(values_only=True)):
        if i == 0: continue
        repo.upsert(Student(str(row[0]), row[1], row[2], row[3]))
    return redirect("/dashboard")

if __name__ == "__main__":
    app.run(debug=True)
